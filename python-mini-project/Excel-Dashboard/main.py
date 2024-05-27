import pandas as pd
import numpy as np
from openpyxl import load_workbook 
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment  
from openpyxl.styles import PatternFill, Font

file_name = 'Excel-Dashboard/Bike_sales.xlsx'
bike_df = pd.read_excel(file_name, sheet_name='bike_buyers')

print(bike_df.head())

#We don't want to mess with our raw data, thus, making a copy of it into a sheet called Working_Sheet.
with pd.ExcelWriter(file_name, 
    engine='openpyxl',
    mode='a',
    if_sheet_exists='replace'
    ) as writer:
    bike_df.to_excel(writer, sheet_name='Working_Sheet', index=False)

#Let's read the working sheet data into our dataframe
df = pd.read_excel(file_name,sheet_name='Working_Sheet')

df.drop_duplicates(keep='first', inplace=True, ignore_index=False)
df['Marital Status'] = df['Marital Status'].replace('M', 'Married').replace('S', 'Single')
df['Gender'] = df['Gender'].replace('M', 'Male').replace('F', 'Female')
df['Age brackets'] = df['Age'].apply(lambda x:'Less than 30' if x <=30 else ('Greater than  55' if x > 55 else '31 to 55' ))
df['Commute Distance'] = df['Commute Distance'].replace('10+ Miles', 'More than 10 Miles')
print(df.head())

#Pivot table 1
#Average Income per Gender based on Purchased Yes or No
avg_gender_income_df = np.round(pd.pivot_table(df,
               values = 'Income',
               index = ['Gender'],
               columns = ['Purchased Bike'],
               aggfunc = np.mean
              ),2)

#Now that we have made all changes in the dataframe, let's load it into the excel file
with pd.ExcelWriter(file_name,#Name of the Workbook
                    engine='openpyxl',#Name of the engine
                    mode='a',#Append mode
                   if_sheet_exists="replace" #Replacing the sheet if it already exists
                   ) as writer:  
    avg_gender_income_df.to_excel(writer, sheet_name='Average_Gender_Income')


# loading workbook and selecting sheet
wb = load_workbook(file_name)
sheet = wb['Average_Gender_Income']

# Bar chart creation
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Average Income by Gender and Purchase Data"
chart1.y_axis.title = 'Gender'
chart1.x_axis.title = 'Income'

#Attach the chart to the worksheet
data1 = Reference(sheet, min_col=2, min_row=1, max_row=3, max_col=3)#Including Headers
cats1 = Reference(sheet, min_col=1, min_row=2, max_row=3)#Not including headers
chart1.add_data(data1, titles_from_data=True)
chart1.dataLabels = DataLabelList() 
chart1.dataLabels.showVal = True
chart1.set_categories(cats1)
chart1.shape = 4
sheet.add_chart(chart1, "A10")
wb.save(file_name)

#Pivot table 2
#Count of people purchasing bike per Age brackets
count_age_brackets_purchase_df = pd.pivot_table(df,
                values = 'ID',
                index = ['Age brackets'],
                columns = ['Purchased Bike'],
                aggfunc = 'count'
                )

#Now that we have made all changes in the dataframe, let's load it into the excel file
with pd.ExcelWriter(file_name,#Name of the Workbook
                    engine='openpyxl',#Name of the engine
                    mode='a',#Append mode
                    if_sheet_exists="replace" #Replacing the sheet if it already exists
                    ) as writer:  
    count_age_brackets_purchase_df.to_excel(writer, sheet_name='Count_Age_Brackets')


# loading workbook and selecting sheet
wb = load_workbook(file_name)
sheet2 = wb['Count_Age_Brackets']

chart2 = LineChart()
chart2.style = 10
chart2.title = "Count of Purchased with Age Brackets"
chart2.y_axis.title = 'Count'
chart2.x_axis.title = 'Age brackets'

data2 = Reference(sheet2, min_col=2, max_col=3 , min_row=1, max_row=4)#Including Headers
cats2 = Reference(sheet2, min_col=1, min_row=2, max_row=4)#Not including headers
chart2.add_data(data2, titles_from_data=True)
chart2.dataLabels = DataLabelList() 
chart2.dataLabels.showVal = True
chart2.set_categories(cats2)
chart2.shape = 4
sheet2.add_chart(chart2, "A10")
wb.save(file_name)


#Pivot table 3
#Count of people purchasing bike based on commute distance
count_commute_distance_purchase_df = pd.pivot_table(df,
                values = ['ID'],
                index = 'Commute Distance',
                columns = 'Purchased Bike',
                aggfunc = 'count'
                )


count_commute_distance_purchase_df = count_commute_distance_purchase_df['ID']

#Now that we have made all changes in the dataframe, let's load it into the excel file
with pd.ExcelWriter(file_name,#Name of the Workbook
                    engine='openpyxl',#Name of the engine
                    mode='a',#Append mode
                    if_sheet_exists="replace" #Replacing the sheet if it already exists
                    ) as writer:  
    
    count_commute_distance_purchase_df.to_excel(writer, sheet_name='Count_Commute_Distance')


# loading workbook and selecting sheet
wb = load_workbook(file_name)
sheet3 = wb['Count_Commute_Distance']

chart3 = LineChart()
chart3.style = 10
chart3.title = "Count of Purchased with Commute Distance"
chart3.y_axis.title = 'Count'
chart3.x_axis.title = 'Commute Distance'

data3 = Reference(sheet3, min_col=2, max_col=3, min_row=1, max_row=6 )#Including Headers
cats3 = Reference(sheet3, min_col=1, min_row=2, max_row=6)#Not including headers
chart3.add_data(data3, titles_from_data=True)
chart3.dataLabels = DataLabelList()
chart3.dataLabels.showVal = True
chart3.set_categories(cats3)
chart3.shape = 4
chart3.width = 31
sheet3.add_chart(chart3, "A10")
wb.save(file_name)

#Creating an empty Dataframe
title_df = pd.DataFrame()
#Now that we have made all changes in the dataframe, let's load it into the excel file
with pd.ExcelWriter(file_name,#Name of the Workbook
                    engine='openpyxl',#Name of the engine
                    mode='a',#Append mode
                   if_sheet_exists="replace" #Replacing the sheet if it already exists
                   ) as writer:  
    title_df.to_excel(writer, sheet_name='Dashboard')

# loading workbook and selecting sheet
wb = load_workbook(file_name)
sheet = wb['Dashboard']
for x in range(1,22):
    sheet.merge_cells('A1:R4')
    
cell = sheet.cell(row=1, column=1)  
cell.value = 'Bike Sales Dashboard'  
cell.alignment = Alignment(horizontal='center', vertical='center')   
cell.font  = Font(b=True, color="F8F8F8",size = 46)
cell.fill = PatternFill("solid", fgColor="2591DB")

# Adding all our pivot charts to the dashboard
sheet.add_chart(chart1,'A5')
sheet.add_chart(chart2,'J5')
chart3.width = 31
sheet.add_chart(chart3,'A20')
wb.save(file_name)